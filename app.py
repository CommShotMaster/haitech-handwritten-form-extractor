"""
Handwritten Form Extractor - Desktop Application
A polished GUI app using Gemini 2.5 Flash to extract handwritten forms into Excel.
Built with CustomTkinter for a modern, professional look.
"""

import os
import sys
import json
import re
import threading
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, END, BOTH, LEFT, RIGHT, X, Y

import customtkinter as ctk
import pandas as pd
from PIL import Image
from dotenv import load_dotenv
from google import genai

# HEIC/HEIF support (iPhone photos)
try:
    import pillow_heif  # type: ignore
    pillow_heif.register_heif_opener()
    HEIC_SUPPORTED = True
except ImportError:
    HEIC_SUPPORTED = False


# ─── Configuration ───────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".webp", ".heic", ".heif")
MODEL_NAME = "gemini-2.5-flash"

from extractor_core import process_images_parallel

FIELD_ORDER = [
    "Name (Chinese)", "Name (English)", "Class",
    "Contact Number", "Food Type"
]

FILETYPES = [
    ("Image files", "*.jpg *.jpeg *.png *.bmp *.webp *.heic *.heif"),
    ("HEIC/HEIF (iPhone)", "*.heic *.heif"),
    ("All files", "*.*"),
]


# ─── Theme ───────────────────────────────────────────────────────────────────

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# Custom colors - Premium Light & Dark Aesthetic
PRIMARY = ("#0EA5E9", "#38BDF8")        
PRIMARY_HOVER = ("#0284C7", "#0EA5E9")
PRIMARY_TEXT = ("#FFFFFF", "#0F172A")
SECONDARY = ("#8B5CF6", "#A78BFA")      
SECONDARY_HOVER = ("#7C3AED", "#8B5CF6")
SUCCESS_CLR = ("#10B981", "#34D399")
SUCCESS_HOVER = ("#059669", "#10B981")
DANGER = ("#F43F5E", "#FB7185")
DANGER_HOVER = ("#E11D48", "#F43F5E")
WARNING = ("#F59E0B", "#FBBF24")
SURFACE = ("#F8FAFC", "#0F172A")        # Main app background
SURFACE_2 = ("#FFFFFF", "#1E293B")      # Cards/Panels
SURFACE_3 = ("#F1F5F9", "#334155")      # Hover/Borders/Secondary cards
TEXT_MAIN = ("#0F172A", "#F8FAFC")
TEXT_DIM = ("#64748B", "#94A3B8")
BORDER_CLR = ("#E2E8F0", "#475569")
BTN_DISABLED = ("#E2E8F0", "#334155")


class FormExtractorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # ─── Window Setup ────────────────────────────────────────────────
        self.title("HaiTech — Handwritten Form Extractor")
        self.geometry("1100x780")
        self.minsize(900, 650)
        self.configure(fg_color=SURFACE)

        # Set window icon
        icon_path = BASE_DIR / "icon.ico"
        if icon_path.exists():
            self.iconbitmap(str(icon_path))

        # State
        self.selected_files = []
        self.all_data = []
        self.client = None
        self.is_processing = False
        self.active_api_key = ""
        self._cancel_event = threading.Event()

        # Init API (soft — won't exit if missing, shows panel instead)
        self._init_api(soft=True)

        # Build UI
        self._build_header()
        self._build_api_section()   # ← new API key panel
        self._build_file_section()
        self._build_action_bar()
        self._build_results_section()
        self._build_footer()

        # Grid weights
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)  # results row shifted by 1
        
        # Maximize window automatically
        self.after(100, lambda: self.state("zoomed"))

    # ─── API Setup ───────────────────────────────────────────────────────────

    def _init_api(self, soft=False):
        """Initialize the Gemini API client from .env file."""
        load_dotenv(BASE_DIR / ".env", override=True)
        api_key = os.getenv("GEMINI_API_KEY", "")

        if not api_key or api_key == "your_api_key_here":
            if not soft:
                messagebox.showerror(
                    "API Key Missing",
                    "Gemini API key not found!\n\n"
                    "Enter your key in the API Key panel below."
                )
            self.client = None
            self.active_api_key = ""
            return

        self.client = genai.Client(api_key=api_key)
        self.active_api_key = api_key

    def _apply_api_key(self):
        """Validate the entered API key and update the client + .env file."""
        raw_key = self.api_key_entry.get().strip()
        if not raw_key:
            messagebox.showwarning("Empty Key", "Please paste your Gemini API key first.")
            return

        # Save to .env file so it persists
        env_path = BASE_DIR / ".env"
        env_lines = []
        key_written = False
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                if line.startswith("GEMINI_API_KEY"):
                    env_lines.append(f"GEMINI_API_KEY={raw_key}")
                    key_written = True
                else:
                    env_lines.append(line)
        if not key_written:
            env_lines.append(f"GEMINI_API_KEY={raw_key}")

        env_path.write_text("\n".join(env_lines) + "\n", encoding="utf-8")

        # Re-init client
        os.environ["GEMINI_API_KEY"] = raw_key
        try:
            self.client = genai.Client(api_key=raw_key)
            self.active_api_key = raw_key
            self._update_api_status(connected=True)
            self._sync_api_edit_state()
            self.api_key_entry.delete(0, "end")
            messagebox.showinfo("API Key Updated", "✅ API key saved and connected successfully!")
        except Exception as e:
            self.client = None
            self.active_api_key = ""
            self._update_api_status(connected=False)
            messagebox.showerror("Connection Error", f"Could not connect with this key:\n{e}")

    def _update_api_status(self, connected: bool):
        """Update the API status indicator in the panel."""
        if connected:
            masked = self.active_api_key[:6] + "••••••••" + self.active_api_key[-4:] if len(self.active_api_key) > 10 else "••••••••"
            self.api_status_label.configure(
                text=f"🟢  Connected  •  Key: {masked}",
                text_color=SUCCESS_CLR
            )
        else:
            self.api_status_label.configure(
                text="🔴  No API key — paste your key and click Apply",
                text_color=DANGER
            )

    def _on_model_change(self, new_model=None):
        """Update header and footer when model changes."""
        if new_model is None:
            if hasattr(self, 'model_var'):
                new_model = self.model_var.get()
            else:
                new_model = "gemini-2.5-flash"
        self.header_subtitle.configure(text=f"Powered by {new_model}  •  HaiTech")
        self.footer_label.configure(text=f"© 2026 HaiTech  •  {new_model}  •  v1.0")

    def _toggle_theme(self):
        """Toggle between Light and Dark mode."""
        mode = self.theme_var.get()
        ctk.set_appearance_mode(mode)

    # ─── UI Building ─────────────────────────────────────────────────────────

    def _build_header(self):
        """Build the top header bar."""
        header = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0)
        header.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 8))

        # Left side: icon + title
        title_frame = ctk.CTkFrame(header, fg_color="transparent")
        title_frame.pack(side=LEFT, fill=X)

        # App icon badge
        icon_badge = ctk.CTkFrame(title_frame, fg_color=PRIMARY, corner_radius=10,
                                   width=44, height=44)
        icon_badge.pack(side=LEFT, padx=(0, 14))
        icon_badge.pack_propagate(False)
        ctk.CTkLabel(icon_badge, text="📝", font=("Segoe UI Emoji", 20)).pack(
            expand=True)

        # Title + subtitle
        text_frame = ctk.CTkFrame(title_frame, fg_color="transparent")
        text_frame.pack(side=LEFT)

        ctk.CTkLabel(
            text_frame, text="Handwritten Form Extractor",
            font=("Segoe UI", 22, "bold"), text_color=TEXT_MAIN
        ).pack(anchor="w")

        self.header_subtitle = ctk.CTkLabel(
            text_frame, text=f"Powered by {MODEL_NAME}  •  HaiTech",
            font=("Segoe UI", 11), text_color=TEXT_DIM
        )
        self.header_subtitle.pack(anchor="w")

        # Right side: Theme toggle
        right_frame = ctk.CTkFrame(header, fg_color="transparent")
        right_frame.pack(side=RIGHT, fill=Y)

        self.theme_var = ctk.StringVar(value=ctk.get_appearance_mode())
        self.theme_switch = ctk.CTkSwitch(
            right_frame, text="🌓", command=self._toggle_theme,
            variable=self.theme_var, onvalue="Dark", offvalue="Light",
            font=("Segoe UI Emoji", 16), fg_color=SURFACE_3, progress_color=PRIMARY
        )
        self.theme_switch.pack(side=RIGHT, pady=10)
        
        if ctk.get_appearance_mode() == "Dark":
            self.theme_switch.select()
        else:
            self.theme_switch.deselect()

        # Separator
        sep = ctk.CTkFrame(self, fg_color=BORDER_CLR, height=1, corner_radius=0)
        sep.grid(row=0, column=0, sticky="sew", padx=24)

    def _build_api_section(self):
        """Build the API key management panel."""
        section = ctk.CTkFrame(self, fg_color=SURFACE_2, corner_radius=14)
        section.grid(row=1, column=0, sticky="ew", padx=24, pady=(6, 0))

        inner = ctk.CTkFrame(section, fg_color="transparent")
        inner.pack(fill=X, padx=16, pady=10)

        # Left: icon + label
        ctk.CTkLabel(
            inner, text="🔑  Gemini API Key",
            font=("Segoe UI", 13, "bold"), text_color=TEXT_MAIN
        ).pack(side=LEFT, padx=(0, 16))

        # Status indicator
        connected = bool(self.active_api_key)
        masked = self.active_api_key[:6] + "••••••••" + self.active_api_key[-4:] if len(self.active_api_key) > 10 else "••••••••"
        status_text = f"🟢  Connected  •  Key: {masked}" if connected else "🔴  No API key — paste your key and click Apply"
        status_color = SUCCESS_CLR if connected else DANGER

        self.api_status_label = ctk.CTkLabel(
            inner, text=status_text,
            font=("Segoe UI", 11), text_color=status_color
        )
        self.api_status_label.pack(side=LEFT, padx=(0, 20))

        # Right: entry + button
        right = ctk.CTkFrame(inner, fg_color="transparent")
        right.pack(side=RIGHT)

        self.api_edit_container = ctk.CTkFrame(right, fg_color="transparent")
        self.api_edit_container.pack(side=LEFT, padx=(0, 8))

        self.edit_api_btn = ctk.CTkButton(
            self.api_edit_container, text="✏️ Edit Key",
            font=("Segoe UI", 11, "bold"), fg_color=SURFACE_3, hover_color=BORDER_CLR,
            text_color=TEXT_MAIN, corner_radius=8, height=34, width=90,
            command=self._toggle_api_edit
        )

        self.api_key_entry = ctk.CTkEntry(
            self.api_edit_container, width=280, height=34,
            placeholder_text="Paste new API key here…",
            font=("Consolas", 11),
            fg_color=SURFACE, border_color=BORDER_CLR,
            show="•"
        )
        
        self.apply_key_btn = ctk.CTkButton(
            self.api_edit_container, text="Apply Key",
            font=("Segoe UI", 12, "bold"),
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER, text_color=PRIMARY_TEXT,
            corner_radius=8, height=34, width=100,
            command=self._apply_api_key
        )

        # Model selector dropdown (ComboBox allows typing custom models)
        self.model_var = ctk.StringVar(value=MODEL_NAME)
        self.model_var.trace_add("write", lambda *args: self._on_model_change())
        self.model_dropdown = ctk.CTkComboBox(
            right, values=[
                "gemini-2.5-flash", "gemini-2.5-pro", 
                "gemini-2.0-flash", "gemini-2.0-pro-exp",
                "gemini-1.5-pro", "gemini-1.5-flash",
                "gemini-exp-1206", "learnlm-1.5-pro-experimental"
            ],
            variable=self.model_var,
            font=("Segoe UI", 11),
            width=160, height=34,
            fg_color=SURFACE_3, border_color=BORDER_CLR, button_color=BORDER_CLR, button_hover_color=PRIMARY,
            command=self._on_model_change
        )
        self.model_dropdown.pack(side=LEFT, padx=(8, 0))
        
        self._sync_api_edit_state()

    def _sync_api_edit_state(self, force_edit=False):
        """Show edit button or entry depending on connection state."""
        self.edit_api_btn.pack_forget()
        self.api_key_entry.pack_forget()
        self.apply_key_btn.pack_forget()

        if bool(self.active_api_key) and not force_edit:
            self.edit_api_btn.pack(side=LEFT)
        else:
            self.api_key_entry.pack(side=LEFT, padx=(0, 8))
            self.apply_key_btn.pack(side=LEFT)

    def _toggle_api_edit(self):
        self._sync_api_edit_state(force_edit=True)

    def _build_file_section(self):
        """Build the file selection area."""
        section = ctk.CTkFrame(self, fg_color=SURFACE_2, corner_radius=14)
        section.grid(row=2, column=0, sticky="ew", padx=24, pady=(6, 6))

        # Header row
        header_row = ctk.CTkFrame(section, fg_color="transparent")
        header_row.pack(fill=X, padx=16, pady=(14, 8))

        ctk.CTkLabel(
            header_row, text="📁  File Selection",
            font=("Segoe UI", 15, "bold"), text_color=TEXT_MAIN
        ).pack(side=LEFT)

        self.file_count_label = ctk.CTkLabel(
            header_row, text="No files selected",
            font=("Segoe UI", 12), text_color=TEXT_DIM
        )
        self.file_count_label.pack(side=LEFT, padx=(14, 0))

        # Buttons row
        btn_row = ctk.CTkFrame(header_row, fg_color="transparent")
        btn_row.pack(side=RIGHT)

        self.select_btn = ctk.CTkButton(
            btn_row, text="📎  Select Images", font=("Segoe UI", 12, "bold"),
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER, text_color=PRIMARY_TEXT,
            text_color_disabled=PRIMARY_TEXT,
            corner_radius=8, height=36, width=150,
            command=self._select_files
        )
        self.select_btn.pack(side=LEFT, padx=(0, 8))

        self.folder_btn = ctk.CTkButton(
            btn_row, text="📂  Select Folder", font=("Segoe UI", 12, "bold"),
            fg_color=SECONDARY, hover_color=SECONDARY_HOVER, text_color=PRIMARY_TEXT,
            text_color_disabled=PRIMARY_TEXT,
            corner_radius=8, height=36, width=150,
            command=self._select_folder
        )
        self.folder_btn.pack(side=LEFT, padx=(0, 8))

        self.clear_btn = ctk.CTkButton(
            btn_row, text="✕  Clear", font=("Segoe UI", 11),
            fg_color=SURFACE_3, hover_color=BORDER_CLR,
            text_color=TEXT_DIM, corner_radius=8, height=36, width=90,
            command=self._clear_files
        )
        self.clear_btn.pack(side=LEFT)

        # File list
        self.file_textbox = ctk.CTkTextbox(
            section, height=100, font=("Consolas", 11),
            fg_color=SURFACE, corner_radius=8,
            border_color=BORDER_CLR, border_width=1,
            text_color=TEXT_MAIN, state="disabled"
        )
        self.file_textbox.pack(fill=X, padx=16, pady=(0, 14))

    def _build_action_bar(self):
        """Build the extract button + progress area."""
        action = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0)
        action.grid(row=3, column=0, sticky="ew", padx=24, pady=(6, 6))

        # Top row: button + status
        top_row = ctk.CTkFrame(action, fg_color="transparent")
        top_row.pack(fill=X)

        self.extract_btn = ctk.CTkButton(
            top_row, text="🚀  Extract All Forms",
            font=("Segoe UI", 14, "bold"),
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER, text_color=PRIMARY_TEXT,
            text_color_disabled=PRIMARY_TEXT,
            corner_radius=10, height=46,
            command=self._start_extraction
        )
        self.extract_btn.pack(side=LEFT, fill=X, expand=True, padx=(0, 12))

        self.review_btn = ctk.CTkButton(
            top_row, text="🔍  Review & Confirm",
            font=("Segoe UI", 13, "bold"),
            fg_color=SECONDARY, hover_color=SECONDARY_HOVER, text_color=PRIMARY_TEXT,
            text_color_disabled=PRIMARY_TEXT,
            corner_radius=10, height=46, width=190,
            command=self._open_review, state="disabled"
        )
        self.review_btn.pack(side=RIGHT, padx=(0, 10))

        self.save_btn = ctk.CTkButton(
            top_row, text="💾  Save to Excel",
            font=("Segoe UI", 13, "bold"),
            fg_color=SUCCESS_CLR, hover_color=SUCCESS_HOVER, text_color=PRIMARY_TEXT,
            text_color_disabled=PRIMARY_TEXT,
            corner_radius=10, height=46, width=170,
            command=self._save_excel, state="disabled"
        )
        self.save_btn.pack(side=RIGHT)

        # Progress bar
        self.progress_bar = ctk.CTkProgressBar(
            action, height=6, corner_radius=3,
            fg_color=SURFACE_3, progress_color=PRIMARY
        )
        self.progress_bar.pack(fill=X, pady=(10, 2))
        self.progress_bar.set(0)

        # Status label
        self.status_label = ctk.CTkLabel(
            action, text="Ready — select images to begin",
            font=("Segoe UI", 11), text_color=TEXT_DIM, anchor="w"
        )
        self.status_label.pack(fill=X, pady=(2, 0))

    def _build_results_section(self):
        """Build the results table area."""
        section = ctk.CTkFrame(self, fg_color=SURFACE_2, corner_radius=14)
        section.grid(row=4, column=0, sticky="nsew", padx=24, pady=(6, 8))

        # Header
        header = ctk.CTkFrame(section, fg_color="transparent")
        header.pack(fill=X, padx=16, pady=(14, 8))

        ctk.CTkLabel(
            header, text="📊  Extraction Results",
            font=("Segoe UI", 15, "bold"), text_color=TEXT_MAIN
        ).pack(side=LEFT)

        self.result_count_label = ctk.CTkLabel(
            header, text="",
            font=("Segoe UI", 12), text_color=TEXT_DIM
        )
        self.result_count_label.pack(side=LEFT, padx=(14, 0))

        # Scrollable results frame
        self.results_scroll = ctk.CTkScrollableFrame(
            section, fg_color=SURFACE, corner_radius=8,
            border_color=BORDER_CLR, border_width=1,
            label_text=""
        )
        self.results_scroll.pack(fill=BOTH, expand=True, padx=16, pady=(0, 14))

        # Column headers
        self._build_table_header()

    def _build_table_header(self):
        """Build the table column headers."""
        header_frame = ctk.CTkFrame(self.results_scroll, fg_color=SURFACE_3,
                                     corner_radius=6, height=36)
        header_frame.pack(fill=X, pady=(0, 4))
        header_frame.grid_propagate(False)

        columns = ["#", "File"] + FIELD_ORDER
        weights = [1, 5, 4, 5, 2, 3, 2]

        for i, (col, w) in enumerate(zip(columns, weights)):
            header_frame.grid_columnconfigure(i, weight=w, uniform="col")
            lbl = ctk.CTkLabel(
                header_frame, text=col, font=("Segoe UI", 12, "bold"),
                text_color=PRIMARY, anchor="w"
            )
            lbl.grid(row=0, column=i, sticky="ew", padx=(8 if i == 0 else 4, 4), pady=4)

    def _build_footer(self):
        """Build the bottom footer."""
        footer = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0)
        footer.grid(row=5, column=0, sticky="ew", padx=24, pady=(0, 10))

        self.footer_label = ctk.CTkLabel(
            footer, text=f"© 2026 HaiTech  •  {MODEL_NAME}  •  v1.0",
            font=("Segoe UI", 10), text_color=TEXT_DIM
        )
        self.footer_label.pack()

    # ─── Actions ─────────────────────────────────────────────────────────────

    def _select_files(self):
        """Open file dialog to select individual images."""
        files = filedialog.askopenfilenames(
            title="Select Handwritten Form Images",
            filetypes=FILETYPES,
            initialdir=str(BASE_DIR / "uploads")
        )
        if files:
            self.selected_files = list(files)
            self._update_file_list()

    def _select_folder(self):
        """Open folder dialog to select all images in a folder."""
        folder = filedialog.askdirectory(
            title="Select Folder Containing Form Images",
            initialdir=str(BASE_DIR / "uploads")
        )
        if folder:
            folder_path = Path(folder)
            found = []
            for f in sorted(folder_path.iterdir()):
                if f.is_file() and f.suffix.lower() in IMAGE_EXTENSIONS:
                    found.append(str(f))
            if not found:
                messagebox.showinfo(
                    "No Images Found",
                    f"No image files found in:\n{folder}"
                )
                return
            self.selected_files = found
            self._update_file_list()

    def _update_file_list(self):
        """Update the file display with current selection."""
        self.file_textbox.configure(state="normal")
        self.file_textbox.delete("1.0", END)
        for i, f in enumerate(self.selected_files, 1):
            self.file_textbox.insert(END, f"  {i}.  {Path(f).name}\n")
        self.file_textbox.configure(state="disabled")

        count = len(self.selected_files)
        self.file_count_label.configure(
            text=f"{count} file{'s' if count != 1 else ''} selected",
            text_color=SUCCESS_CLR
        )
        self.status_label.configure(
            text=f"✅ {count} image(s) ready — click Extract to begin",
            text_color=SUCCESS_CLR
        )

    def _clear_files(self):
        """Clear all selected files."""
        self.selected_files = []
        self.file_textbox.configure(state="normal")
        self.file_textbox.delete("1.0", END)
        self.file_textbox.configure(state="disabled")
        self.file_count_label.configure(text="No files selected", text_color=TEXT_DIM)
        self.status_label.configure(text="Ready — select images to begin", text_color=TEXT_DIM)
        self.progress_bar.set(0)

    def _start_extraction(self):
        """Start the extraction process in a background thread."""
        if not self.client:
            messagebox.showerror(
                "No API Key",
                "Please enter and apply a Gemini API key first.\n\n"
                "Get a free key at:\nhttps://aistudio.google.com/apikey"
            )
            return

        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select image files first.")
            return

        if self.is_processing:
            return

        self.is_processing = True
        self._cancel_event.clear()

        # Update extract button to act as a cancel button
        self.extract_btn.configure(
            text="🛑  Cancel Extraction",
            fg_color=DANGER, hover_color=DANGER_HOVER,
            command=self._cancel_extraction
        )
        
        # Disable buttons
        self.select_btn.configure(state="disabled")
        self.folder_btn.configure(state="disabled")
        self.save_btn.configure(state="disabled")
        self.review_btn.configure(state="disabled")

        # Clear previous results
        for widget in self.results_scroll.winfo_children()[1:]:  # Keep header
            widget.destroy()
        self.all_data = []
        self.progress_bar.set(0)

        # Run in background
        thread = threading.Thread(target=self._run_extraction, daemon=True)
        thread.start()

    def _cancel_extraction(self):
        """Cancel the running extraction."""
        if self.is_processing:
            self.status_label.configure(text="🛑  Cancelling...", text_color=DANGER)
            self._cancel_event.set()
            self.extract_btn.configure(state="disabled", text="⏳  Cancelling...")

    def _run_extraction(self):
        """Extract data from all selected images using parallel processing (background thread)."""
        total = len(self.selected_files)

        def _progress_cb(completed, total_files, filename, data):
            # Update progress
            self.after(0, lambda v=completed/total_files: self.progress_bar.set(v))
            # Add row
            self.after(0, lambda d=data, idx=completed: self._add_result_row(d, idx))
            # Update status label
            self.after(0, lambda fn=filename, c=completed: self.status_label.configure(
                text=f"📄  Processed ({c}/{total_files}): {fn}",
                text_color=WARNING
            ))

        self.all_data = process_images_parallel(
            client=self.client,
            filepaths=self.selected_files,
            model_name=self.model_var.get(),
            progress_callback=_progress_cb,
            cancel_event=self._cancel_event,
            max_workers=3
        )

        self.after(0, self._extraction_done)

    def _add_result_row(self, data, row_num):
        """Add a result row to the table."""
        is_error = data.get("_status") == "error"
        row_color = SURFACE_2 if row_num % 2 == 0 else SURFACE

        row_frame = ctk.CTkFrame(self.results_scroll, fg_color=row_color,
                                  corner_radius=4, height=32)
        row_frame.pack(fill=X, pady=1)
        row_frame.grid_propagate(False)

        weights = [1, 5, 4, 5, 2, 3, 2]
        for i, w in enumerate(weights):
            row_frame.grid_columnconfigure(i, weight=w, uniform="col")

        # Row number
        ctk.CTkLabel(
            row_frame, text=str(row_num), font=("Segoe UI", 12),
            text_color=TEXT_DIM, anchor="w"
        ).grid(row=0, column=0, sticky="ew", padx=(8, 4), pady=2)

        # Filename
        ctk.CTkLabel(
            row_frame, text=data.get("_source_file", "")[:45],
            font=("Segoe UI", 12), text_color=TEXT_MAIN, anchor="w"
        ).grid(row=0, column=1, sticky="ew", padx=4, pady=2)

        if is_error:
            ctk.CTkLabel(
                row_frame,
                text=f"❌ {data.get('_error', 'Error')}",
                font=("Segoe UI", 12), text_color=DANGER,
                anchor="w"
            ).grid(row=0, column=2, columnspan=5, sticky="ew", padx=4, pady=2)
        else:
            for i, field in enumerate(FIELD_ORDER):
                value = str(data.get(field, ""))
                ctk.CTkLabel(
                    row_frame, text=value[:45] if len(value) > 45 else value,
                    font=("Segoe UI", 12), text_color=TEXT_MAIN, anchor="w"
                ).grid(row=0, column=i + 2, sticky="ew", padx=4, pady=2)

    def _extraction_done(self):
        """Called when extraction is complete."""
        total = len(self.all_data)
        success = sum(1 for d in self.all_data if d.get("_status") == "success")

        if self._cancel_event.is_set():
            self.status_label.configure(
                text=f"🛑  Extraction Cancelled. Extracted {success} of {total} form(s)",
                text_color=DANGER
            )
        else:
            self.status_label.configure(
                text=f"✅  Done! Successfully extracted {success} of {total} form(s)",
                text_color=SUCCESS_CLR
            )
            
        self.result_count_label.configure(
            text=f"{success}/{total} successful"
        )
        self.progress_bar.set(1.0)

        # Restore extract button
        self.extract_btn.configure(
            state="normal", 
            text="🚀  Extract All Forms",
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER, text_color=PRIMARY_TEXT,
            command=self._start_extraction
        )
        self.select_btn.configure(state="normal")
        self.folder_btn.configure(state="normal")
        
        has_data = len(self.all_data) > 0
        self.save_btn.configure(state="normal" if has_data else "disabled")
        self.review_btn.configure(state="normal" if has_data else "disabled")
        self.is_processing = False

    def _open_review(self):
        """Open the Review & Confirm dialog."""
        if not self.all_data:
            messagebox.showwarning("No Data", "Please extract forms first.")
            return
        dialog = ReviewDialog(
            parent=self,
            all_data=self.all_data,
            selected_files=self.selected_files,
            on_data_confirmed=lambda updated: self._apply_reviewed_data(updated)
        )
        dialog.focus()

    def _apply_reviewed_data(self, updated_data):
        """Apply confirmed edits from the review dialog back to the app."""
        self.all_data = updated_data
        # Refresh the table view
        for widget in self.results_scroll.winfo_children()[1:]:
            widget.destroy()
        for i, data in enumerate(self.all_data, 1):
            self._add_result_row(data, i)
        self.status_label.configure(
            text="✅  Review complete — data updated.",
            text_color=SUCCESS_CLR
        )
        # Automatically trigger the actual Excel save
        self._save_excel()

    def _save_excel(self):
        """Save results to Excel file with custom format."""
        if not self.all_data:
            messagebox.showwarning("No Data", "No extraction results to save.")
            return

        # 1. Ask for Event Title
        dialog = ctk.CTkInputDialog(text="Enter the Event Title:\n(This will be the sheet title and default file name)", title="Event Title")
        event_title = dialog.get_input()
        
        if not event_title:
            return  # Cancelled

        # Format a safe filename
        safe_title = "".join([c if c.isalnum() or c in " -_" else "_" for c in event_title])
        default_name = f"{safe_title}.xlsx"

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        filepath = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=str(OUTPUT_DIR),
            initialfile=default_name
        )

        if not filepath:
            return

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Side, Font

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Forms"

            # Define styles
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            bold_font = Font(bold=True)
            title_font = Font(bold=True, size=14)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            # 1. Add Title Row (Merged A1:J1)
            ws.merge_cells('A1:J1')
            ws['A1'] = event_title
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align

            # 2. Add Headers (Row 2)
            headers = ["", "Label", "Attd.", "NAME", "NAME", "Class", "Meal", "Mobile", "", "REMARK"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border

            # 3. Add Data
            current_row = 3
            for data in self.all_data:
                if data.get("_status") == "success":
                    # Fill row with empty strings first to apply borders
                    for col_idx in range(1, 11):
                        cell = ws.cell(row=current_row, column=col_idx, value="")
                        cell.border = thin_border
                    
                    ws.cell(row=current_row, column=4, value=data.get("Name (Chinese)", "")).alignment = left_align
                    ws.cell(row=current_row, column=5, value=data.get("Name (English)", "")).alignment = left_align
                    ws.cell(row=current_row, column=6, value=data.get("Class", "")).alignment = center_align
                    ws.cell(row=current_row, column=7, value=data.get("Food Type", "")).alignment = center_align
                    ws.cell(row=current_row, column=8, value=data.get("Contact Number", "")).alignment = left_align
                    
                    current_row += 1

            # 4. Adjust Column Widths
            col_widths = {
                'A': 5, 'B': 8, 'C': 8, 
                'D': 15, 'E': 30, 'F': 10, 
                'G': 10, 'H': 18, 'I': 18, 'J': 20
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            wb.save(filepath)

            self.status_label.configure(
                text=f"💾  Saved to: {Path(filepath).name}",
                text_color=SUCCESS_CLR
            )
            messagebox.showinfo("Saved!", f"Excel file saved to:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save file:\n{e}")


# ─── Review Dialog ──────────────────────────────────────────────────────────

class ReviewDialog(ctk.CTkToplevel):
    """Side-by-side image + editable fields review window."""

    def __init__(self, parent, all_data, selected_files, on_data_confirmed):
        super().__init__(parent)
        self.title("🔍  Review & Confirm Extracted Forms")
        self.geometry("1100x680")
        self.minsize(900, 600)
        self.configure(fg_color=SURFACE)
        
        # Maximize window automatically for better visibility on small screens
        self.after(100, lambda: self.state("zoomed"))

        # Deep-copy data so edits don't affect original until confirmed
        self.all_data = [dict(d) for d in all_data]
        self.selected_files = selected_files
        self.on_data_confirmed = on_data_confirmed
        self.current_index = 0
        self.confirmed = [False] * len(self.all_data)
        self.field_vars = {}
        self.current_rotation = -90  # Default to rotated right (portrait -> landscape)
        self._current_ctk_image = None  # prevent GC
        self.target_img_w = 600
        self.target_img_h = 600
        self.zoom_factor = 1.0
        self.pan_x = 0
        self.pan_y = 0
        self.drag_start_x = 0
        self.drag_start_y = 0
        self._resize_timer = None

        self._build_ui()
        self._load_entry(self.current_index)
        self.grab_set()
        self.focus()

    # ── Build UI ─────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Top bar ──
        top = ctk.CTkFrame(self, fg_color=SURFACE_2, corner_radius=0, height=54)
        top.pack(fill=X)
        top.pack_propagate(False)

        ctk.CTkLabel(
            top, text="🔍  Review & Confirm Extracted Forms",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_MAIN
        ).pack(side=LEFT, padx=20, pady=14)

        self.counter_label = ctk.CTkLabel(
            top, text="", font=("Segoe UI", 12), text_color=TEXT_DIM
        )
        self.counter_label.pack(side=LEFT, padx=8)

        self.confirmed_badge = ctk.CTkLabel(
            top, text="", font=("Segoe UI", 11, "bold"), text_color=SUCCESS_CLR
        )
        self.confirmed_badge.pack(side=LEFT, padx=8)

        # ── Main content: left image | right fields ──
        content = ctk.CTkFrame(self, fg_color="transparent")
        # We delay packing `content` to prioritize `nav` in the layout engine!
        # uniform="colgroup" prevents the layout from shifting when rotating!
        content.grid_columnconfigure(0, weight=60, uniform="colgroup")
        content.grid_columnconfigure(1, weight=40, uniform="colgroup")
        content.grid_rowconfigure(0, weight=1)

        # Left image panel
        img_panel = ctk.CTkFrame(content, fg_color=SURFACE_2, corner_radius=12)
        img_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        img_top = ctk.CTkFrame(img_panel, fg_color="transparent")
        img_top.pack(fill=X, padx=14, pady=(12, 6))
        ctk.CTkLabel(
            img_top, text="📷  Original Form",
            font=("Segoe UI", 13, "bold"), text_color=TEXT_MAIN
        ).pack(side=LEFT)
        self.img_name_label = ctk.CTkLabel(
            img_top, text="", font=("Segoe UI", 10), text_color=TEXT_DIM
        )
        self.img_name_label.pack(side=LEFT, padx=(10, 0))

        # Rotate button
        ctk.CTkButton(
            img_top, text="Rotate 🔄",
            font=("Segoe UI", 11, "bold"),
            fg_color=SURFACE_3, hover_color=BORDER_CLR,
            text_color=TEXT_MAIN, corner_radius=6, height=24, width=80,
            command=self._rotate_image
        ).pack(side=RIGHT)

        self.image_label = ctk.CTkLabel(
            img_panel, text="", fg_color=SURFACE, corner_radius=8
        )
        self.image_label.pack(fill=BOTH, expand=True, padx=12, pady=(0, 12))
        self.image_label.bind("<Configure>", self._on_resize)
        self.image_label.bind("<MouseWheel>", self._on_mousewheel)
        self.image_label.bind("<ButtonPress-1>", self._on_drag_start)
        self.image_label.bind("<B1-Motion>", self._on_drag_motion)

        # Right fields panel
        right_panel = ctk.CTkFrame(content, fg_color=SURFACE_2, corner_radius=12)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        ctk.CTkLabel(
            right_panel, text="📝  Extracted Information  (editable)",
            font=("Segoe UI", 13, "bold"), text_color=TEXT_MAIN
        ).pack(anchor="w", padx=16, pady=(14, 6))

        ctk.CTkLabel(
            right_panel,
            text="Edit any field below, click 💾 Update to save changes, then ✅ Confirm.",
            font=("Segoe UI", 10), text_color=TEXT_DIM
        ).pack(anchor="w", padx=16, pady=(0, 8))

        fields_scroll = ctk.CTkScrollableFrame(
            right_panel, fg_color="transparent", corner_radius=0
        )
        fields_scroll.pack(fill=BOTH, expand=True, padx=12, pady=(0, 8))
        fields_scroll.grid_columnconfigure(0, weight=1)

        self.field_vars = {}
        for field in FIELD_ORDER:
            card = ctk.CTkFrame(fields_scroll, fg_color=SURFACE_3, corner_radius=8)
            card.pack(fill=X, pady=5, padx=2)

            ctk.CTkLabel(
                card, text=field,
                font=("Segoe UI", 13, "bold"), text_color=TEXT_DIM, anchor="w"
            ).pack(fill=X, padx=16, pady=(14, 6))

            var = ctk.StringVar()
            entry = ctk.CTkEntry(
                card, textvariable=var,
                font=("Segoe UI", 18),
                fg_color=SURFACE, border_color=BORDER_CLR,
                height=54
            )
            entry.pack(fill=X, padx=16, pady=(0, 16))
            self.field_vars[field] = var

        # ── Bottom nav bar ──
        nav = ctk.CTkFrame(self, fg_color=SURFACE_2, corner_radius=0, height=68)
        nav.pack(side="bottom", fill=X, pady=(10, 0))
        nav.pack_propagate(False)
        
        # Now pack content so it takes the remaining space
        content.pack(side="top", fill=BOTH, expand=True, padx=14, pady=(10, 0))

        # Prev / Next
        nav_left = ctk.CTkFrame(nav, fg_color="transparent")
        nav_left.pack(side=LEFT, padx=16, pady=14)

        self.prev_btn = ctk.CTkButton(
            nav_left, text="◀  Prev",
            font=("Segoe UI", 12, "bold"),
            fg_color=SURFACE_3, hover_color=BORDER_CLR,
            text_color=TEXT_MAIN, corner_radius=8, height=40, width=110,
            command=self._go_prev
        )
        self.prev_btn.pack(side=LEFT, padx=(0, 8))

        self.next_btn = ctk.CTkButton(
            nav_left, text="Next  ▶",
            font=("Segoe UI", 12, "bold"),
            fg_color=SURFACE_3, hover_color=BORDER_CLR,
            text_color=TEXT_MAIN, corner_radius=8, height=40, width=110,
            command=self._go_next
        )
        self.next_btn.pack(side=LEFT)

        # Centre: Update + Confirm
        nav_mid = ctk.CTkFrame(nav, fg_color="transparent")
        nav_mid.pack(side=LEFT, expand=True, padx=20, pady=14)

        self.update_btn = ctk.CTkButton(
            nav_mid, text="💾  Update Edits",
            font=("Segoe UI", 12, "bold"),
            fg_color=SECONDARY, hover_color=SECONDARY_HOVER, text_color=PRIMARY_TEXT,
            corner_radius=8, height=40, width=160,
            command=self._update_current
        )
        self.update_btn.pack(side=LEFT, padx=(0, 10))

        self.confirm_btn = ctk.CTkButton(
            nav_mid, text="✅  Confirm This Form",
            font=("Segoe UI", 13, "bold"),
            fg_color=SUCCESS_CLR, hover_color=SUCCESS_HOVER, text_color=PRIMARY_TEXT,
            corner_radius=8, height=40, width=200,
            command=self._confirm_current
        )
        self.confirm_btn.pack(side=LEFT)

        # Right: Save All
        nav_right = ctk.CTkFrame(nav, fg_color="transparent")
        nav_right.pack(side=RIGHT, padx=16, pady=14)

        ctk.CTkButton(
            nav_right, text="💾  Save All to Excel",
            font=("Segoe UI", 12, "bold"),
            fg_color=PRIMARY, hover_color=PRIMARY_HOVER, text_color=PRIMARY_TEXT,
            corner_radius=8, height=40, width=180,
            command=self._save_and_close
        ).pack()

    # ── Logic ────────────────────────────────────────────────────────────

    def _load_entry(self, idx):
        """Load form image + field values for the given index."""
        data = self.all_data[idx]
        total = len(self.all_data)
        confirmed_count = sum(self.confirmed)

        # Counter & badge
        self.counter_label.configure(text=f"Form {idx + 1} of {total}")
        self.confirmed_badge.configure(
            text=f"✅  {confirmed_count} confirmed" if confirmed_count else ""
        )

        # Nav buttons
        self.prev_btn.configure(state="normal" if idx > 0 else "disabled")
        self.next_btn.configure(state="normal" if idx < total - 1 else "disabled")

        # Confirm button
        if self.confirmed[idx]:
            self.confirm_btn.configure(
                text="✅  Already Confirmed",
                fg_color=BTN_DISABLED, state="disabled", text_color=TEXT_DIM
            )
        else:
            self.confirm_btn.configure(
                text="✅  Confirm This Form",
                fg_color=SUCCESS_CLR, state="normal"
            )

        # Load image
        filepath = self.selected_files[idx] if idx < len(self.selected_files) else None
        self.img_name_label.configure(text=data.get("_source_file", ""))

        if filepath and Path(filepath).exists():
            try:
                self.current_pil_image = Image.open(filepath)
                self.zoom_factor = 1.0
                self.pan_x = 0
                self.pan_y = 0
                self._render_image()
            except Exception as e:
                self.current_pil_image = None
                self._current_ctk_image = None
                self.image_label.configure(
                    image=None, text=f"❌ Could not load image\n{e}",
                    text_color=DANGER, font=("Segoe UI", 11)
                )
        else:
            self.current_pil_image = None
            self._current_ctk_image = None
            self.image_label.configure(
                image=None, text="No image available",
                text_color=TEXT_DIM, font=("Segoe UI", 12)
            )

        # Populate fields
        for field in FIELD_ORDER:
            val = data.get(field, "")
            self.field_vars[field].set(str(val) if val else "")

    def _rotate_image(self):
        """Rotate the currently displayed image by 90 degrees."""
        if not getattr(self, 'current_pil_image', None):
            return
        self.current_rotation = (self.current_rotation - 90) % 360
        self._render_image()

    def _on_resize(self, event):
        """Dynamically track the available space for the image."""
        if event.width > 50 and event.height > 50:
            if abs(event.width - getattr(self, 'target_img_w', 600)) > 20 or abs(event.height - getattr(self, 'target_img_h', 600)) > 20:
                self.target_img_w = event.width
                self.target_img_h = event.height
                if getattr(self, '_resize_timer', None):
                    self.after_cancel(self._resize_timer)
                self._resize_timer = self.after(150, self._render_image)

    def _on_mousewheel(self, event):
        """Handle zooming in and out."""
        if not getattr(self, 'current_pil_image', None):
            return
        if event.delta > 0:
            self.zoom_factor *= 1.1
        elif event.delta < 0:
            self.zoom_factor /= 1.1
        self.zoom_factor = max(0.5, min(self.zoom_factor, 5.0))
        self._render_image()

    def _on_drag_start(self, event):
        """Record the start position for panning."""
        self.drag_start_x = event.x
        self.drag_start_y = event.y

    def _on_drag_motion(self, event):
        """Handle panning the image."""
        if not getattr(self, 'current_pil_image', None):
            return
        dx = event.x - self.drag_start_x
        dy = event.y - self.drag_start_y
        self.drag_start_x = event.x
        self.drag_start_y = event.y
        self.pan_x += dx
        self.pan_y += dy
        self._render_image()

    def _render_image(self):
        """Render the current PIL image with rotation, zoom, and pan applied."""
        if not getattr(self, 'current_pil_image', None):
            return
            
        img = self.current_pil_image.copy()
        if self.current_rotation != 0:
            img = img.rotate(self.current_rotation, expand=True)
            
        target_w = max(100, getattr(self, 'target_img_w', 600) - 4)
        target_h = max(100, getattr(self, 'target_img_h', 600) - 4)
        
        # Calculate fit scale
        scale_w = target_w / img.width
        scale_h = target_h / img.height
        base_scale = min(scale_w, scale_h)
        
        # Apply user zoom
        final_scale = base_scale * self.zoom_factor
        new_w = int(img.width * final_scale)
        new_h = int(img.height * final_scale)
        
        # Resize image
        img = img.resize((new_w, new_h), Image.LANCZOS)
        
        # Create a background canvas of target size
        bg_color = (248, 250, 252) if ctk.get_appearance_mode() == "Light" else (15, 23, 42)
        bg = Image.new("RGB", (target_w, target_h), bg_color)
        
        # Center with pan
        paste_x = (target_w - new_w) // 2 + self.pan_x
        paste_y = (target_h - new_h) // 2 + self.pan_y
        
        bg.paste(img, (paste_x, paste_y))
        
        ctk_img = ctk.CTkImage(
            light_image=bg, dark_image=bg, size=(target_w, target_h)
        )
        self._current_ctk_image = ctk_img
        self.image_label.configure(image=ctk_img, text="")

    def _flush_edits(self):
        """Write current field values back into all_data for the current form."""
        has_data = False
        for field in FIELD_ORDER:
            val = self.field_vars[field].get().strip()
            self.all_data[self.current_index][field] = val
            if val:
                has_data = True
                
        # If the user manually filled in data for a form that previously failed,
        # clear the error status so it saves properly to Excel!
        if has_data and self.all_data[self.current_index].get("_status") != "success":
            self.all_data[self.current_index]["_status"] = "success"
            if "_error" in self.all_data[self.current_index]:
                del self.all_data[self.current_index]["_error"]

    def _go_prev(self):
        self._flush_edits()
        self.current_index -= 1
        self._load_entry(self.current_index)

    def _go_next(self):
        self._flush_edits()
        self.current_index += 1
        self._load_entry(self.current_index)

    def _update_current(self):
        """Save edits immediately with visual feedback, without confirming."""
        self._flush_edits()
        # Flash the button green to confirm save
        self.update_btn.configure(text="✅  Updated!", fg_color=SUCCESS_CLR)
        self.after(1200, lambda: self.update_btn.configure(
            text="💾  Update Edits", fg_color=SECONDARY
        ))

    def _confirm_current(self):
        self._flush_edits()
        self.confirmed[self.current_index] = True
        self._load_entry(self.current_index)

        # Auto-jump to next unconfirmed form
        for i, c in enumerate(self.confirmed):
            if not c:
                self.current_index = i
                self._load_entry(self.current_index)
                return

    def _save_and_close(self):
        """Flush any unsaved edits, pass data back to main app, close."""
        self._flush_edits()
        self.on_data_confirmed(self.all_data)
        self.destroy()


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    app = FormExtractorApp()
    app.mainloop()


if __name__ == "__main__":
    main()
